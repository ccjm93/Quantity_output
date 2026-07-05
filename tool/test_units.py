# -*- coding: utf-8 -*-
"""Excel/COM 없이 실행 가능한 순수 로직 유닛테스트.

실행: python tool/test_units.py
(pywin32/pymupdf/openpyxl 은 import 용으로만 필요하며 Excel 실행은 필요 없다.)
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import annotate  # noqa: E402
import detect  # noqa: E402
import review  # noqa: E402
import structure  # noqa: E402
import toolruntime  # noqa: E402
from config import (  # noqa: E402
    compile_exclude_patterns,
    load_settings,
    normalize_pdf_path,
)


class TestStructure(unittest.TestCase):
    def test_natural_key_order(self):
        names = ["10. 부대공", "2. 토공", "1. 공통공사"]
        self.assertEqual(sorted(names, key=structure.natural_key),
                         ["1. 공통공사", "2. 토공", "10. 부대공"])

    def test_skip_names(self):
        self.assertTrue(structure._is_skip("~$임시.xlsx", "_output", "_backup"))
        self.assertTrue(structure._is_skip("_output", "_output", "_backup"))
        self.assertTrue(structure._is_skip("루트_backup_20260101", "_output", "_backup"))
        self.assertFalse(structure._is_skip("2.0 토공.xlsx", "_output", "_backup"))

    def test_scan_skips_tool_artifacts(self):
        with tempfile.TemporaryDirectory() as d:
            open(os.path.join(d, "1. 공통.xlsx"), "w").close()
            open(os.path.join(d, "출력물 오류 검토결과.xlsx"), "w").close()
            events = structure.scan(d, [".xlsx"],
                                    skip_names={"출력물 오류 검토결과.xlsx"})
            self.assertEqual([e["name"] for e in events], ["1. 공통.xlsx"])

    def test_scan_orders_folders_and_files(self):
        with tempfile.TemporaryDirectory() as d:
            os.makedirs(os.path.join(d, "2. 취수틀"))
            os.makedirs(os.path.join(d, "10. 기타"))
            open(os.path.join(d, "1. 공통.xlsx"), "w").close()
            open(os.path.join(d, "2. 취수틀", "2.0 토공.xlsx"), "w").close()
            open(os.path.join(d, "10. 기타", "9.0 부대공.xlsx"), "w").close()
            events = structure.scan(d, [".xlsx"])
            names = [e["name"] for e in events]
            self.assertEqual(names, ["1. 공통.xlsx", "2. 취수틀", "2.0 토공.xlsx",
                                     "10. 기타", "9.0 부대공.xlsx"])

    def test_scan_prunes_folders_without_files(self):
        """엑셀 파일이 없는 폴더는 간지(빈 페이지)만 남으므로 이벤트에서 제외한다."""
        with tempfile.TemporaryDirectory() as d:
            os.makedirs(os.path.join(d, "3. 빈폴더"))
            os.makedirs(os.path.join(d, "4. 중첩", "4.1 빈하위"))  # 파일 없는 중첩 폴더
            open(os.path.join(d, "1. 공통.xlsx"), "w").close()
            events = structure.scan(d, [".xlsx"])
            names = [e["name"] for e in events]
            self.assertEqual(names, ["1. 공통.xlsx"])


class TestConfig(unittest.TestCase):
    def test_defaults_loaded(self):
        cfg = load_settings(None)
        self.assertEqual(cfg["output_dir_name"], "_output")
        self.assertFalse(cfg["set_print_area_if_missing"])
        self.assertIn("exclude_sheet_patterns", cfg)

    def test_exclude_patterns_match(self):
        pats = compile_exclude_patterns(load_settings(None)["exclude_sheet_patterns"])

        def excluded(name):
            return any(p.search(name) for p in pats)

        self.assertTrue(excluded("이후시트는 출력하지마세요"))
        self.assertTrue(excluded("토공집계 (2)"))
        self.assertTrue(excluded("Sheet1"))
        self.assertFalse(excluded("주요자재집계표"))


class TestDetectRegex(unittest.TestCase):
    def test_hash_only(self):
        self.assertTrue(detect._HASH_ONLY.match("####"))
        self.assertTrue(detect._HASH_ONLY.match("##  "))
        self.assertFalse(detect._HASH_ONLY.match("#"))
        self.assertFalse(detect._HASH_ONLY.match("#REF!"))
        self.assertFalse(detect._HASH_ONLY.match("가##"))


class TestReview(unittest.TestCase):
    def test_text_findings_ref_and_hash(self):
        found = review._text_findings("합계 #REF! 오류 ####", 3, {})
        types = [f["types"][0] for f in found]
        self.assertIn("#REF! 오류", types)
        self.assertIn("# 연속 표시", types)

    def test_groups(self):
        self.assertEqual(review._groups([1, 2, 3, 10, 11, 30]),
                         [(1, 3), (10, 11), (30, 30)])
        self.assertEqual(review._groups([]), [])


class TestToolruntime(unittest.TestCase):
    def test_normalize_pdf_path(self):
        self.assertTrue(normalize_pdf_path("out").endswith(".pdf"))
        self.assertTrue(normalize_pdf_path("out.PDF").endswith("out.PDF"))

    def test_korean_error_types(self):
        k = toolruntime._korean_error_types
        self.assertEqual(k(["#REF! 오류"]), "#REF! 오류")
        self.assertEqual(k(["# 연속 표시"]), "# 연속 표시")
        self.assertEqual(k(["수식 오류"]), "수식 오류")
        self.assertEqual(k(["빈 페이지"]), "빈 페이지 의심")
        self.assertEqual(k(["하단 표 테두리 누락 의심"]), "하단 표 테두리 누락 의심")
        self.assertEqual(k(["#REF! 오류", "#REF! 오류"]), "#REF! 오류")  # 중복 제거

    def test_assign_pages(self):
        manifest = [
            {"type": "file_divider", "page": 1, "file": "a.xlsx", "sheet": None},
            {"type": "sheet", "page": 2, "file": "a.xlsx", "sheet": "토공"},
            {"type": "sheet", "page": 3, "file": "a.xlsx", "sheet": "토공"},
        ]
        findings = [{"file": "a.xlsx", "sheet": "토공", "cell": "B2",
                     "kind": "overflow", "text": "###"}]
        toolruntime._assign_pages(findings, manifest)
        self.assertEqual(findings[0]["page"], 2)  # 시트 첫 페이지
        self.assertEqual(findings[0]["types"], ["# 연속 표시"])

    def test_find_review_pdf_prefers_configured_name(self):
        with tempfile.TemporaryDirectory() as d:
            open(os.path.join(d, "AAA 다른문서.pdf"), "w").close()
            open(os.path.join(d, "수량산출서 output.pdf"), "w").close()
            cfg = load_settings(None)
            pdf = toolruntime._find_review_pdf(d, cfg)
            self.assertEqual(os.path.basename(pdf), "수량산출서 output.pdf")

    def test_find_review_pdf_ignores_marked_copy(self):
        """검토 표시 사본만 있는 폴더에서는 그것을 검토 대상으로 삼지 않는다."""
        with tempfile.TemporaryDirectory() as d:
            open(os.path.join(d, "다른이름 output(검토표시).pdf"), "w").close()
            cfg = load_settings(None)
            self.assertIsNone(toolruntime._find_review_pdf(d, cfg))

    def test_display_width(self):
        self.assertEqual(toolruntime._display_width("abc"), 3)
        self.assertEqual(toolruntime._display_width("가나"), 4)

    def test_rotate_backups_keeps_recent(self):
        with tempfile.TemporaryDirectory() as d:
            root = os.path.join(d, "루트")
            os.makedirs(root)
            for ts in ("20260101_000000", "20260102_000000",
                       "20260103_000000", "20260104_000000"):
                os.makedirs(os.path.join(d, f"루트_backup_{ts}"))
            toolruntime._rotate_backups(root, "_backup", 2, log=lambda *a: None)
            remain = sorted(n for n in os.listdir(d) if "_backup_" in n)
            self.assertEqual(remain, ["루트_backup_20260103_000000",
                                      "루트_backup_20260104_000000"])

    def test_prune_old_reports_keeps_recent(self):
        with tempfile.TemporaryDirectory() as d:
            for ts in ("20260101_000000", "20260102_000000", "20260103_000000"):
                open(os.path.join(d, f"review_{ts}.json"), "w").close()
                open(os.path.join(d, f"cell_issues_{ts}.json"), "w").close()
            toolruntime._prune_old_reports(d, 1, log=lambda *a: None)
            remain = sorted(os.listdir(d))
            self.assertEqual(remain, ["cell_issues_20260103_000000.json",
                                      "review_20260103_000000.json"])

    def test_load_cell_findings_uses_latest(self):
        import json
        with tempfile.TemporaryDirectory() as d:
            old = {"findings": [{"cell": "A1", "kind": "error"}]}
            new = {"findings": [{"cell": "B2", "kind": "overflow"},
                                {"cell": "C3", "kind": "ref_error"}]}
            with open(os.path.join(d, "cell_issues_20260101_000000.json"),
                      "w", encoding="utf-8") as f:
                json.dump(old, f)
            with open(os.path.join(d, "cell_issues_20260102_000000.json"),
                      "w", encoding="utf-8") as f:
                json.dump(new, f)
            found = toolruntime._load_cell_findings(d, log=lambda *a: None)
            self.assertEqual([x["cell"] for x in found], ["B2", "C3"])  # 최신 파일

    def test_load_cell_findings_empty_dir(self):
        with tempfile.TemporaryDirectory() as d:
            self.assertEqual(toolruntime._load_cell_findings(d, log=lambda *a: None), [])

    def test_save_review_xlsx_without_excel(self):
        rev = {"findings": [{"page": 5, "types": ["#REF! 오류"],
                             "detail": "시트!B2 → '#REF!'", "file": "a.xlsx",
                             "folder": "", "sheet": "토공"},
                            {"page": 7, "types": ["하단 표 테두리 누락 의심"],
                             "detail": "하단 테두리 의심", "file": "b.xlsx",
                             "folder": "", "sheet": "집계"}]}
        with tempfile.TemporaryDirectory() as d:
            path = toolruntime.save_review_xlsx(rev, d, log=lambda *a: None)
            self.assertTrue(os.path.isfile(path))
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(ws["C2"].value, "#REF! 오류")
            self.assertEqual(ws["B2"].value, "최종 PDF 5페이지")
            self.assertEqual(ws.freeze_panes, "A2")
            # 확실한 오류(빨강) / 의심(주황) 행 색상
            self.assertEqual(ws["C2"].fill.fgColor.rgb, "00FFC7CE")
            self.assertEqual(ws["C3"].fill.fgColor.rgb, "00FFEB9C")

    def test_row_severity(self):
        self.assertEqual(toolruntime._row_severity("#REF! 오류"), "error")
        self.assertEqual(toolruntime._row_severity("수식 오류"), "error")
        self.assertEqual(toolruntime._row_severity("# 연속 표시"), "error")
        self.assertEqual(toolruntime._row_severity("하단 표 테두리 누락 의심"), "warn")
        self.assertEqual(toolruntime._row_severity("빈 페이지 의심"), "warn")
        self.assertEqual(toolruntime._row_severity("이상 없음"), "ok")


class TestAnnotate(unittest.TestCase):
    def test_marked_pdf_path(self):
        self.assertEqual(annotate.marked_pdf_path(r"C:\x\out.pdf"),
                         r"C:\x\out(검토표시).pdf")

    def test_search_text_priority(self):
        self.assertEqual(annotate._search_text({"text": "######"}), "######")
        self.assertEqual(annotate._search_text(
            {"detail": "연속 # 표시가 발견되었습니다: '####'"}), "####")
        self.assertEqual(annotate._search_text({"types": ["#REF! 오류"]}), "#REF!")
        self.assertIsNone(annotate._search_text({"types": ["기타"]}))

    def test_create_marked_pdf_draws_boxes(self):
        import fitz
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "src.pdf")
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "#REF!")
            doc.save(src)
            doc.close()
            findings = [
                {"page": 1, "types": ["#REF! 오류"], "detail": "x → '#REF!'"},
                {"page": 1, "types": ["하단 표 테두리 누락 의심"]},
                {"page": 99, "types": ["#REF! 오류"]},  # 범위 밖 → 무시
            ]
            out = os.path.join(d, "out.pdf")
            path = annotate.create_marked_pdf(src, findings, out, log=lambda *a: None)
            self.assertEqual(path, out)
            marked = fitz.open(out)
            drawings = marked.load_page(0).get_drawings()
            marked.close()
            self.assertEqual(len(drawings), 2)  # 텍스트 박스 1 + 하단 영역 박스 1

    def test_create_marked_pdf_dedupes_same_text(self):
        """같은 페이지에서 같은 텍스트를 가리키는 finding 여러 개 → 박스는 한 번만."""
        import fitz
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "src.pdf")
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "#DIV/0!")
            doc.save(src)
            doc.close()
            findings = [
                {"page": 1, "types": ["수식 오류"], "text": "#DIV/0!"},
                {"page": 1, "types": ["수식 오류"], "text": "#DIV/0!"},
                {"page": 1, "types": ["수식 오류"], "text": "#DIV/0!"},
            ]
            out = os.path.join(d, "out.pdf")
            annotate.create_marked_pdf(src, findings, out, log=lambda *a: None)
            marked = fitz.open(out)
            drawings = marked.load_page(0).get_drawings()
            marked.close()
            self.assertEqual(len(drawings), 1)

    def test_create_marked_pdf_searches_all_sheet_pages(self):
        """excel finding 은 manifest 상 해당 시트의 모든 페이지에서 텍스트를 찾는다."""
        import fitz
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "src.pdf")
            doc = fitz.open()
            doc.new_page()                                # p.1: 시트 첫 페이지(오류 없음)
            doc.new_page().insert_text((72, 72), "####")  # p.2: 같은 시트 둘째 페이지
            doc.save(src)
            doc.close()
            manifest = [
                {"type": "sheet", "page": 1, "file": "a.xlsx", "sheet": "토공"},
                {"type": "sheet", "page": 2, "file": "a.xlsx", "sheet": "토공"},
            ]
            findings = [{"page": 1, "source": "excel", "file": "a.xlsx",
                         "sheet": "토공", "types": ["# 연속 표시"], "text": "####"}]
            out = os.path.join(d, "out.pdf")
            annotate.create_marked_pdf(src, findings, out,
                                       manifest=manifest, log=lambda *a: None)
            marked = fitz.open(out)
            self.assertEqual(len(marked.load_page(0).get_drawings()), 0)
            self.assertEqual(len(marked.load_page(1).get_drawings()), 1)
            marked.close()

    def test_create_marked_pdf_no_findings(self):
        self.assertIsNone(annotate.create_marked_pdf(
            "unused.pdf", [{"page": None}], "out.pdf", log=lambda *a: None))


if __name__ == "__main__":
    unittest.main(verbosity=2)
