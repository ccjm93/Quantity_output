# -*- coding: utf-8 -*-
"""수량산출서 출력 자동화 도구 - CLI / 오케스트레이션 (단일 PDF 병합판).

흐름: 백업 → 구조 스캔 → 파일별(PBP 활성화+원본저장 → 시트별 임시 PDF)
      → 간지 포함 단일 PDF 병합 → 프로그램 내부 오류 검토 → 산출물 저장.

사용 예:
  python tool/toolruntime.py "01 수량_가야"
  python tool/toolruntime.py "루트" --no-backup       # 원본 백업 생략(주의)
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import shutil
import signal
import sys
import tempfile
import traceback

try:
    # line_buffering: exe(frozen) 자식 프로세스에는 -u 를 못 넘기므로
    # GUI 로그가 실시간 스트리밍되도록 줄 단위 플러시를 보장한다.
    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    sys.stderr.reconfigure(encoding="utf-8", line_buffering=True)
except Exception:
    pass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

import annotate as annotate_mod  # noqa: E402
import detect as detect_mod  # noqa: E402
import export as export_mod  # noqa: E402
import prep as prep_mod  # noqa: E402
import review as review_mod  # noqa: E402
import structure as structure_mod  # noqa: E402
from config import (  # noqa: E402
    compile_exclude_patterns,
    load_settings,
    normalize_pdf_path,
)
from excel_app import ExcelApp  # noqa: E402
from merge import Merger  # noqa: E402

REVIEW_XLSX_NAME = "출력물 오류 검토결과.xlsx"


def _assign_pages(cell_findings, manifest):
    """각 (file, sheet) 의 통합 PDF 첫 페이지를 구해 cell finding 에 page 를 부여한다."""
    first_page = {}
    for m in manifest:
        if m.get("type") == "sheet":
            key = (m.get("file"), m.get("sheet"))
            p = m.get("page")
            if key not in first_page or p < first_page[key]:
                first_page[key] = p
    for f in cell_findings:
        f["page"] = first_page.get((f.get("file"), f.get("sheet")))
        f["source"] = "excel"
        if f["kind"] == "ref_error":
            f["types"] = ["#REF! 오류"]
        elif f["kind"] == "error":
            f["types"] = ["수식 오류"]
        else:
            f["types"] = ["# 연속 표시"]
        f["detail"] = f"{f.get('sheet')}!{f.get('cell')} → '{f.get('text')}'"
    return cell_findings


def _original_location(finding):
    file_path = (finding.get("file") or "").replace("\\", "/").strip("/")
    folder = (finding.get("folder") or "").replace("\\", "/").strip("/")
    if file_path:
        return "/" + file_path
    if folder:
        return "/" + folder
    return ""


def _display_location(finding):
    page = finding.get("page")
    return f"최종 PDF {page}페이지" if page else ""


# 파이프라인이 생성하는 오류 유형 → 결과표 표기. 매핑에 없으면 원문 그대로 표기.
_ERROR_TYPE_LABELS = {"빈 페이지": "빈 페이지 의심"}

# 결과표 행 색상: 확실한 오류(빨강) / 휴리스틱 의심(주황) / 이상 없음(초록)
_CERTAIN_TYPES = ("#REF! 오류", "수식 오류", "# 연속 표시")
_ROW_STYLES = {
    "error": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")),
    "warn": (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C6500")),
    "ok": (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")),
}


def _row_severity(types_text: str) -> str:
    if types_text == "이상 없음":
        return "ok"
    if any(t in types_text for t in _CERTAIN_TYPES):
        return "error"
    return "warn"


def _korean_error_types(types):
    if isinstance(types, str):
        types = [types]
    out = []
    for t in types or []:
        raw = str(t).strip()
        if raw:
            out.append(_ERROR_TYPE_LABELS.get(raw, raw))
    return ", ".join(dict.fromkeys(out))


def _display_width(text):
    """열 폭 산정용 표시 폭(한글 등 전각 문자는 2로 계산)."""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(text))


def save_review_xlsx(review_result, out_dir, log=print):
    """프로그램 내부 오류 검토 결과를 사용자가 바로 열 수 있는 xlsx 표로 저장한다.

    openpyxl 로 직접 작성하므로 Excel(COM) 없이도 동작한다(--review 단독 실행 포함)."""
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, REVIEW_XLSX_NAME)
    headers = ["검토방식", "출력물 위치", "오류형태", "오류내용", "원본위치"]
    rows = []
    method = "프로그램 내부 로직"
    for fnd in review_result.get("findings", []):
        rows.append([
            method,
            _display_location(fnd),
            _korean_error_types(fnd.get("types", [])),
            str(fnd.get("detail", "") or ""),
            _original_location(fnd),
        ])
    if not rows:
        rows.append([method, "", "이상 없음", "검토 결과 발견된 오류가 없습니다.", ""])

    wb = Workbook()
    ws = wb.active
    ws.title = "출력물 오류 검토결과"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r, row in enumerate(rows, start=2):
        fill, font = _ROW_STYLES[_row_severity(row[2])]
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.fill = fill
            cell.font = font
    for col in range(1, len(headers) + 1):
        widths = [_display_width(headers[col - 1])]
        widths += [_display_width(row[col - 1]) for row in rows]
        ws.column_dimensions[get_column_letter(col)].width = min(80, max(widths) + 3)
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
    ws.freeze_panes = "A2"
    wb.save(path)
    log(f"출력물 오류 검토결과 Excel: {path}")
    return path


def _rotate_backups(root, suffix, keep, log):
    """root 옆의 '<root명><suffix>_<시각>' 백업 폴더 중 최근 keep개만 남긴다(0=무제한)."""
    if not keep or keep <= 0:
        return
    base = os.path.basename(root.rstrip("\\/")) + suffix + "_"
    parent = os.path.dirname(root.rstrip("\\/")) or "."
    try:
        olds = sorted(
            n for n in os.listdir(parent)
            if n.startswith(base) and os.path.isdir(os.path.join(parent, n)))
    except Exception:
        return
    for name in olds[:-keep]:
        shutil.rmtree(os.path.join(parent, name), ignore_errors=True)
        log(f"오래된 백업 삭제(최근 {keep}개 유지): {name}")


def _prune_old_reports(out_dir, keep, log):
    """_output 안의 시각 스탬프 리포트를 종류별로 최근 keep개만 남긴다(0=무제한)."""
    if not keep or keep <= 0:
        return
    for prefix in ("review_", "cell_issues_"):
        try:
            names = sorted(
                n for n in os.listdir(out_dir)
                if n.startswith(prefix) and n.endswith(".json"))
        except Exception:
            return
        for name in names[:-keep]:
            try:
                os.remove(os.path.join(out_dir, name))
            except Exception:
                pass


def backup_root(root, suffix, log, output_dir_name="_output", keep=3):
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = root.rstrip("\\/") + f"{suffix}_{ts}"
    log(f"원본 백업 생성 중: {dst}")
    # 백업 목적은 '원본 엑셀' 보호 — 이전 실행이 남긴 산출물은 백업에서 제외한다.
    shutil.copytree(root, dst,
                    ignore=shutil.ignore_patterns(
                        output_dir_name, "*_backup_*", "~$*",
                        "manifest.json", "process_report.json",
                        "cell_issues_*.json", "review_*.json",
                        REVIEW_XLSX_NAME, "*.pdf"))
    _rotate_backups(root, suffix, keep, log)
    return dst


def run(root, no_backup=False, out_path=None, log=print) -> int:
    root = os.path.abspath(root)
    if not os.path.isdir(root):
        log(f"[오류] 폴더가 아닙니다: {root}")
        return 2

    cfg = load_settings(root)
    patterns = compile_exclude_patterns(cfg["exclude_sheet_patterns"])

    # 출력 위치: 사용자가 '출력 PDF'를 지정하면 그 폴더에 모든 산출물(PDF/manifest/리포트)을 둔다.
    # 기본은 루트\<output_dir_name>\ — 원본 폴더를 어지럽히지 않고, 백업/재실행 시 누적도 막는다.
    if out_path:
        out_path = normalize_pdf_path(out_path)
        out_root = os.path.dirname(out_path)
    else:
        out_root = os.path.join(root, cfg.get("output_dir_name", "_output"))
        out_path = os.path.join(out_root, cfg.get("output_pdf_name", "수량산출서 output.pdf"))

    log("출력물 오류 검토: 프로그램 내부 로직 사용")

    # [0] 백업
    if cfg.get("backup", True) and not no_backup:
        try:
            backup_root(root, cfg.get("backup_suffix", "_backup"), log,
                        cfg.get("output_dir_name", "_output"),
                        cfg.get("backup_keep", 3))
        except Exception as e:
            log(f"[오류] 백업 실패로 중단(원본 보호): {e}")
            return 3
    else:
        log("원본 백업 생략됨(주의: 원본이 변경됩니다).")

    # [1] 구조 스캔
    events = structure_mod.scan(root, cfg["extensions"],
                                cfg["output_dir_name"], cfg.get("backup_suffix", "_backup"),
                                skip_names={REVIEW_XLSX_NAME})
    log("구조 파악: " + structure_mod.summarize(events))
    if not any(e["type"] == "file" for e in events):
        log("[경고] 처리할 엑셀 파일이 없습니다.")
        return 1

    merger = Merger(cfg.get("divider_font"), log=log)
    tmp_root = tempfile.mkdtemp(prefix="qto_")
    file_no = 0
    n_files = sum(1 for e in events if e["type"] == "file")
    report = {"root": root, "files": [], "errors": []}
    detect_on = cfg.get("detect_cell_issues", True)
    cell_findings = []  # [{file, sheet, cell, kind, text}] — Excel 단계 결정론 탐지

    # [2] COM 세션에서 파일별 처리
    with ExcelApp() as xl:
        app = xl.app
        for ev in events:
            if ev["type"] == "folder":
                merger.add_folder_divider(ev["name"], ev["relpath"], ev["depth"])
                continue
            file_no += 1
            log(f"[{file_no}/{n_files}] {ev['relpath']}")
            try:
                wb = app.Workbooks.Open(os.path.abspath(ev["abspath"]),
                                        UpdateLinks=0, ReadOnly=False)
            except Exception as e:
                log(f"    [오류] 열기 실패, 건너뜀: {e}")
                report["errors"].append({"file": ev["relpath"], "error": str(e)})
                continue
            try:
                pr = prep_mod.prep_workbook(app, wb, patterns,
                                            cfg.get("set_print_area_if_missing", False))
                # 셀 표시 오류(###/수식오류) 결정론적 탐지 (워크북 열린 상태에서)
                if detect_on:
                    before_n = len(cell_findings)  # 이번 파일 분량 집계용
                    for sname in pr["included"]:
                        try:
                            issues, truncated = detect_mod.scan_sheet_issues(
                                wb.Worksheets(sname))
                        except Exception as e:
                            log(f"    [경고] 셀 탐지 실패({sname}): {e}")
                            continue
                        if truncated:
                            log(f"    [경고] 셀 탐지 상한 도달({sname}): "
                                "시트가 매우 커서 일부 셀은 검사되지 않았습니다.")
                        for it in issues:
                            it["file"] = ev["relpath"]
                            cell_findings.append(it)
                    new = cell_findings[before_n:]
                    n_ov = sum(1 for f in new if f["kind"] == "overflow")
                    n_ref = sum(1 for f in new if f["kind"] == "ref_error")
                    n_er = sum(1 for f in new if f["kind"] == "error")
                    if n_ov or n_ref or n_er:
                        log(f"    [탐지] 연속 # {n_ov}건, #REF! {n_ref}건, 수식오류 {n_er}건")
                tmp_dir = os.path.join(tmp_root, f"f{file_no:04d}")
                sheets = export_mod.export_sheets(wb, pr["included"], tmp_dir, log=log)
                if sheets:
                    merger.add_file_divider(ev["name"], ev["relpath"])
                    folder_rel = os.path.dirname(ev["relpath"])
                    for s in sheets:
                        merger.add_sheet(s["pdf"], folder_rel, ev["relpath"], s["sheet"])
                    log(f"    -> {len(sheets)}시트 추가 (제외 {len(pr['excluded'])})")
                else:
                    log(f"    -> 포함 시트 없음(제외 {len(pr['excluded'])})")
                report["files"].append({"file": ev["relpath"],
                                        "included": [s["sheet"] for s in sheets],
                                        "excluded": pr["excluded"]})
            except Exception as e:
                log(f"    [오류] 처리 실패, 건너뜀: {e}")
                report["errors"].append({"file": ev["relpath"],
                                         "error": str(e),
                                         "trace": traceback.format_exc()})
            finally:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass

    # [3] 단일 PDF 저장
    if merger.page_count() == 0:
        log("[경고] 생성할 페이지가 없습니다.")
        return 1
    log(f"단일 PDF 저장 중: {out_path} ({merger.page_count()}페이지)")
    manifest = list(merger.manifest)
    merger.save(out_path)

    os.makedirs(out_root, exist_ok=True)
    with open(os.path.join(out_root, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    with open(os.path.join(out_root, "process_report.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    try:
        shutil.rmtree(tmp_root, ignore_errors=True)
    except Exception:
        pass

    # [3.5] Excel 단계 셀 표시 오류 — PDF 생성과 함께 항상 보고/저장
    _assign_pages(cell_findings, manifest)
    n_ov = sum(1 for f in cell_findings if f["kind"] == "overflow")
    n_ref = sum(1 for f in cell_findings if f["kind"] == "ref_error")
    n_er = sum(1 for f in cell_findings if f["kind"] == "error")
    log("")
    log("===== PDF 생성 완료 =====")
    log(f"단일 PDF: {out_path}")
    # 같은 실행의 산출물(cell_issues/review)이 짝으로 묶이도록 스탬프를 공유한다.
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    if cell_findings:
        ci_path = os.path.join(out_root, f"cell_issues_{stamp}.json")
        with open(ci_path, "w", encoding="utf-8") as f:
            json.dump({"overflow": n_ov, "ref_error": n_ref, "error": n_er,
                       "findings": cell_findings}, f, ensure_ascii=False, indent=2)
        log(f"셀 표시 오류: 연속 # {n_ov}건, #REF! {n_ref}건, 수식오류 {n_er}건  → {ci_path}")
        for fnd in cell_findings[:15]:
            log(f"  p.{fnd.get('page')} | {fnd.get('file')} > {fnd.get('sheet')}"
                f"!{fnd.get('cell')} | {fnd['types'][0]} '{fnd.get('text')}'")
    else:
        log("셀 표시 오류(연속 #/#REF!/수식오류): 발견 없음")
    log("  ※ 셀 값이 정상이어도 PDF 하단 표 테두리 누락 의심은 출력물 검토 단계에서 확인합니다.")

    # [4] 출력물 오류 검토 — 프로그램 내부 로직으로 수행
    rev = review_mod.review(out_path, manifest, cfg, log=log, extra_findings=cell_findings)
    rev_path = os.path.join(out_root, f"review_{stamp}.json")
    with open(rev_path, "w", encoding="utf-8") as f:
        json.dump(rev, f, ensure_ascii=False, indent=2)
    xlsx_path = save_review_xlsx(rev, out_root, log=log)
    _prune_old_reports(out_root, cfg.get("report_keep", 10), log)
    _save_marked_pdf(out_path, rev["findings"], cfg, log, manifest=manifest)
    log(f"이상 의심: {rev['issue_count']}건 ({rev['mode']})  → {rev_path}")
    log(f"검토 결과표: {xlsx_path}")
    for fnd in rev["findings"][:10]:
        log(f"  p.{fnd['page']} | {fnd.get('file')} > {fnd.get('sheet')} "
            f"| {','.join(fnd.get('types', []))} {fnd.get('detail','')}")
    return 0


def _save_marked_pdf(pdf, findings, cfg, log, manifest=None):
    """검토 결과를 빨간 표시로 그린 PDF 사본을 만든다(원본 PDF는 그대로 둠)."""
    if not cfg.get("marked_pdf", True) or not findings:
        return None
    if annotate_mod.MARKED_SUFFIX in os.path.basename(pdf):
        return None  # 표시본을 다시 검토한 경우 — 이중 표시 방지
    try:
        return annotate_mod.create_marked_pdf(
            pdf, findings, annotate_mod.marked_pdf_path(pdf),
            manifest=manifest, font_path=cfg.get("divider_font"), log=log)
    except Exception as e:
        log(f"[경고] 검토 표시 PDF 생성 실패: {e}")
        return None


def _find_review_pdf(folder, cfg):
    """폴더에서 검토 대상 통합 PDF를 찾는다. 설정된 출력 파일명을 최우선으로 매칭."""
    exact = os.path.join(folder, cfg.get("output_pdf_name", "수량산출서 output.pdf"))
    if os.path.isfile(exact):
        return exact
    try:
        # 검토 표시 사본((검토표시).pdf)은 검토 대상에서 제외
        cands = [f for f in os.listdir(folder)
                 if f.lower().endswith(".pdf") and annotate_mod.MARKED_SUFFIX not in f]
    except Exception:
        return None
    if not cands:
        return None
    merged = [f for f in cands if "output" in f.lower() or "통합" in f] or cands

    def _mtime(name):
        try:
            return os.path.getmtime(os.path.join(folder, name))
        except OSError:
            return 0
    # 후보가 여럿이면 가장 최근에 생성/수정된 PDF를 검토 대상으로 삼는다.
    return os.path.join(folder, max(merged, key=_mtime))


def _load_cell_findings(out_dir, log):
    """PDF 생성 시 저장된 최신 cell_issues_*.json 을 읽어 Excel 단계 탐지 결과를 재사용한다.

    검토 전용 실행은 엑셀을 열지 않으므로 셀 단위 탐지를 새로 할 수 없다. 대신
    생성 단계가 남긴 결과(page/types/detail 이미 매핑됨)를 병합해, [PDF 생성 시작]과
    [출력물 오류 검토]의 결과표가 항상 같은 내용이 되도록 한다."""
    try:
        names = sorted(n for n in os.listdir(out_dir)
                       if n.startswith("cell_issues_") and n.endswith(".json"))
    except Exception:
        return []
    if not names:
        return []
    path = os.path.join(out_dir, names[-1])
    try:
        with open(path, encoding="utf-8") as f:
            findings = json.load(f).get("findings", [])
    except Exception as e:
        log(f"[경고] 셀 표시 오류 결과 읽기 실패({names[-1]}): {e}")
        return []
    if findings:
        log(f"PDF 생성 시 탐지된 셀 표시 오류 병합: {len(findings)}건 ({names[-1]})")
    return findings


def run_review_only(target, log=print) -> int:
    """이미 생성된 통합 PDF에 대해 프로그램 내부 오류 검토만 수행."""
    target = os.path.abspath(target)
    if os.path.isdir(target):
        cfg = load_settings(target)
        # 지정 폴더 → 그 아래 산출물 폴더(_output) 순으로 통합 PDF를 찾는다.
        pdf = _find_review_pdf(target, cfg)
        if not pdf:
            sub = os.path.join(target, cfg.get("output_dir_name", "_output"))
            if os.path.isdir(sub):
                pdf = _find_review_pdf(sub, cfg)
        if not pdf:
            log(f"[오류] 검토할 PDF를 찾을 수 없습니다: {target}")
            return 2
        out_dir = os.path.dirname(pdf)
    else:
        pdf = target
        out_dir = os.path.dirname(pdf)
        # 설정은 PDF가 있는 폴더 → 그 상위(루트) 순으로 찾는다.
        # (_output\output.pdf 를 지정해도 루트의 qto_settings.json 이 적용되도록)
        cfg_dir = out_dir
        if not os.path.isfile(os.path.join(cfg_dir, "qto_settings.json")):
            parent = os.path.dirname(out_dir)
            if os.path.isfile(os.path.join(parent, "qto_settings.json")):
                cfg_dir = parent
        cfg = load_settings(cfg_dir)
    if not os.path.isfile(pdf):
        log(f"[오류] PDF 없음: {pdf}")
        return 2

    manifest = []
    mpath = os.path.join(out_dir, "manifest.json")
    if os.path.isfile(mpath):
        try:
            with open(mpath, encoding="utf-8") as f:
                manifest = json.load(f)
        except Exception:
            pass
    else:
        log("[안내] manifest.json 이 없어 위치(폴더/파일/시트) 매핑이 제한됩니다.")

    log(f"출력물 오류 검토 대상: {pdf}")
    cell_findings = _load_cell_findings(out_dir, log)
    rev = review_mod.review(pdf, manifest, cfg, log=log, extra_findings=cell_findings)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    rev_path = os.path.join(out_dir, f"review_{stamp}.json")
    with open(rev_path, "w", encoding="utf-8") as f:
        json.dump(rev, f, ensure_ascii=False, indent=2)
    xlsx_path = save_review_xlsx(rev, out_dir, log=log)
    _prune_old_reports(out_dir, cfg.get("report_keep", 10), log)
    _save_marked_pdf(pdf, rev["findings"], cfg, log, manifest=manifest)
    log("")
    log("===== 검토 완료 =====")
    log(f"이상 의심: {rev['issue_count']}건 ({rev['mode']}) → {rev_path}")
    log(f"검토 결과표: {xlsx_path}")
    for fnd in rev["findings"][:15]:
        log(f"  p.{fnd['page']} | {fnd.get('file')} > {fnd.get('sheet')} "
            f"| {','.join(fnd.get('types', []))} {fnd.get('detail','')}")
    return 0


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        prog="수량산출서 출력 자동화 도구",
        description="루트 폴더의 모든 시트를 폴더·파일 순서로 단일 PDF(간지 포함)로 병합합니다.")
    p.add_argument("input", nargs="?", help="수량산출서 루트 폴더")
    p.add_argument("--out", help="출력 PDF 경로(기본: 루트\\_output\\수량산출서 output.pdf)")
    p.add_argument("--review", help="기존 통합 PDF(또는 출력 폴더)에 대해 오류 검토만 수행")
    p.add_argument("--no-backup", action="store_true", help="원본 백업 생략(주의)")
    args = p.parse_args(argv)

    # GUI가 CTRL_BREAK 로 중단 요청하면 KeyboardInterrupt 로 바꿔
    # with 블록(ExcelApp)이 정리되도록 한다(고아 EXCEL.EXE 방지).
    def _graceful_abort(signum, frame):
        raise KeyboardInterrupt

    try:
        signal.signal(signal.SIGBREAK, _graceful_abort)
    except (AttributeError, ValueError):
        pass  # 비 Windows 또는 서브스레드 — 무시

    try:
        if args.review:
            return run_review_only(args.review)
        if not args.input:
            p.print_help()
            return 2
        return run(args.input, no_backup=args.no_backup, out_path=args.out)
    except KeyboardInterrupt:
        print("[중단] 사용자 요청으로 작업이 중단되었습니다. (Excel 인스턴스는 정리됨)")
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
